"""Convert bounded OHDSI WhiteRabbit scan reports into source-schema YAML."""

from __future__ import annotations

import re
from io import BytesIO
from zipfile import BadZipFile, ZipFile, is_zipfile

import yaml
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict, Field

from .contracts import SourceSchemaDocument

FIELD_OVERVIEW_SHEET = "Field Overview"
TABLE_OVERVIEW_SHEET = "Table Overview"
MAXIMUM_ARCHIVE_ENTRIES = 500
MAXIMUM_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAXIMUM_COMPRESSION_RATIO = 200
MAXIMUM_TABLES = 250
MAXIMUM_FIELDS = 10_000
MAXIMUM_CELL_CHARACTERS = 10_000
MAXIMUM_GENERATED_YAML_BYTES = 750 * 1024
MAXIMUM_GENERATED_RESPONSE_BYTES = 900_000
MAXIMUM_MODEL_IDENTIFIER_CHARACTERS = 250
MAXIMUM_FIELD_IDENTIFIER_CHARACTERS = 255

_HEADER_ALIASES = {
    "table": "table",
    "field": "field",
    "description": "description",
    "type": "type",
    "max length": "max_length",
    "n rows": "row_count",
    "n rows checked": "rows_checked",
    "fraction empty": "fraction_empty",
    "n unique values": "unique_values",
    "fraction unique": "fraction_unique",
    "average": "average",
    "standard deviation": "standard_deviation",
    "min": "minimum",
    "25%": "percentile_25",
    "median": "median",
    "75%": "percentile_75",
    "max": "maximum",
    "n fields": "field_count",
    "n fields empty": "empty_field_count",
}
_REQUIRED_FIELD_HEADERS = {"table", "field", "type"}


class WhiteRabbitReportError(ValueError):
    """Raised when a workbook is unsafe or is not a supported scan report."""


class WhiteRabbitFieldProfile(BaseModel):
    """Non-value aggregate scan metrics for one normalized source field."""

    model_config = ConfigDict(extra="forbid")

    name: str
    source_field_name: str
    raw_data_type: str
    max_length: int | None = None
    row_count: int | None = None
    rows_checked: int | None = None
    fraction_empty: float | None = None
    unique_values: int | None = None
    unique_values_is_upper_bound: bool = False
    fraction_unique: float | None = None
    candidate_key: bool = False


class WhiteRabbitTableProfile(BaseModel):
    """Safe table and field metrics stored beside a generated YAML revision."""

    model_config = ConfigDict(extra="forbid")

    model_name: str
    source_table_name: str
    description: str = ""
    row_count: int | None = None
    rows_checked: int | None = None
    field_count: int
    empty_field_count: int | None = None
    fields: list[WhiteRabbitFieldProfile]


class WhiteRabbitGeneratedSchema(BaseModel):
    """One generated source-schema document and safe preview metadata."""

    model_config = ConfigDict(extra="forbid")

    file_name: str
    content: str
    source_table_name: str
    model_name: str
    field_count: int
    candidate_key_fields: list[str] = Field(default_factory=list)
    profile: WhiteRabbitTableProfile


class WhiteRabbitImportResult(BaseModel):
    """Deterministic conversion result returned by the private API."""

    model_config = ConfigDict(extra="forbid")

    report_version: str | None = None
    table_count: int
    field_count: int
    renamed_table_count: int
    renamed_field_count: int
    tables: list[WhiteRabbitGeneratedSchema]
    warnings: list[str] = Field(default_factory=list)


def _bounded_text(value: object, *, label: str) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > MAXIMUM_CELL_CHARACTERS:
        raise WhiteRabbitReportError(f"{label} exceeds the cell-size limit.")
    return text


def _normalise_identifier(
    value: object,
    *,
    label: str,
    maximum_characters: int,
    numeric_prefix: str,
) -> str:
    source = _bounded_text(value, label=label)
    if not source:
        raise WhiteRabbitReportError(f"{label} is blank.")
    identifier = re.sub(r"[^A-Za-z0-9_]+", "_", source)
    identifier = re.sub(r"_+", "_", identifier).strip("_").lower()
    if not identifier:
        raise WhiteRabbitReportError(f"{label} cannot form a safe identifier.")
    if identifier[0].isdigit():
        identifier = f"{numeric_prefix}_{identifier}"
    if len(identifier) > maximum_characters:
        raise WhiteRabbitReportError(f"{label} exceeds the identifier limit.")
    return identifier


def _header_key(value: object) -> str | None:
    header = " ".join(_bounded_text(value, label="Header").lower().split())
    return _HEADER_ALIASES.get(header)


def _header_indexes(row: tuple[object, ...]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for index, value in enumerate(row):
        key = _header_key(value)
        if key and key not in indexes:
            indexes[key] = index
    return indexes


def _value(row: tuple[object, ...], indexes: dict[str, int], key: str) -> object:
    index = indexes.get(key)
    return row[index] if index is not None and index < len(row) else None


def _optional_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", "")
    match = re.fullmatch(r"(?:<=|<|>=|>)?\s*(\d+(?:\.0+)?)", text)
    return int(float(match.group(1))) if match else None


def _optional_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def _nonnegative_int(value: object) -> int | None:
    result = _optional_int(value)
    return result if result is not None and result >= 0 else None


def _fraction(value: object) -> float | None:
    result = _optional_float(value)
    return result if result is not None and 0 <= result <= 1 else None


def _unique_values(value: object) -> tuple[int | None, bool]:
    text = _bounded_text(value, label="Unique-value count")
    return _nonnegative_int(value), text.startswith(("<", ">"))


def _canonical_data_type(value: object) -> tuple[str, str]:
    raw = _bounded_text(value, label="Source datatype")
    lowered = raw.lower()
    if any(token in lowered for token in ("char", "text", "string", "clob")):
        return "string", raw
    if "timestamp" in lowered or "datetime" in lowered:
        return "datetime", raw
    if "date" in lowered and "time" not in lowered:
        return "date", raw
    if lowered in {"time", "time without time zone", "time with time zone"}:
        return "time", raw
    if any(token in lowered for token in ("bool", "bit")):
        return "boolean", raw
    if any(token in lowered for token in ("int", "serial")):
        return "integer", raw
    if any(
        token in lowered
        for token in ("number", "numeric", "decimal", "float", "double", "real")
    ):
        return "numeric", raw
    if any(token in lowered for token in ("binary", "blob", "byte")):
        return "binary", raw
    return lowered or "unknown", raw


def _validate_archive(content: bytes) -> None:
    if not content or not is_zipfile(BytesIO(content)):
        raise WhiteRabbitReportError("Choose a valid .xlsx WhiteRabbit report.")
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAXIMUM_ARCHIVE_ENTRIES:
                raise WhiteRabbitReportError("Workbook contains too many archive entries.")
            total_uncompressed = 0
            for entry in entries:
                normalised_path = entry.filename.replace("\\", "/")
                if normalised_path.startswith("/") or ".." in normalised_path.split("/"):
                    raise WhiteRabbitReportError("Workbook contains an unsafe archive path.")
                lowered = normalised_path.lower()
                if entry.flag_bits & 0x1:
                    raise WhiteRabbitReportError(
                        "Encrypted workbook entries are not accepted."
                    )
                if "vbaproject.bin" in lowered or lowered.startswith("xl/externallinks/"):
                    raise WhiteRabbitReportError(
                        "Macro-enabled or externally linked workbooks are not accepted."
                    )
                total_uncompressed += entry.file_size
                if total_uncompressed > MAXIMUM_UNCOMPRESSED_BYTES:
                    raise WhiteRabbitReportError("Workbook expands beyond the safety limit.")
                if (
                    entry.compress_size > 0
                    and entry.file_size / entry.compress_size > MAXIMUM_COMPRESSION_RATIO
                ):
                    raise WhiteRabbitReportError("Workbook compression ratio is unsafe.")
    except BadZipFile as error:
        raise WhiteRabbitReportError("Choose a valid .xlsx WhiteRabbit report.") from error


def _report_version(workbook) -> str | None:
    if "_" not in workbook.sheetnames:
        return None
    sheet = workbook["_"]
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 100), values_only=True):
        cells = [_bounded_text(value, label="WhiteRabbit metadata") for value in row]
        for index, cell in enumerate(cells[:-1]):
            if cell.lower() in {"version", "whiterabbit version"} and cells[index + 1]:
                return cells[index + 1][:100]
        combined = " ".join(cell for cell in cells if cell)
        match = re.search(r"WhiteRabbit(?:\s+version)?\s*[:=]?\s*(\d+(?:\.\d+)+)", combined, re.I)
        if match:
            return match.group(1)
    return None


def _table_overview(workbook) -> tuple[dict[str, dict[str, object]], list[str]]:
    if TABLE_OVERVIEW_SHEET not in workbook.sheetnames:
        return {}, ["Table Overview was not present; table summaries use field-level data."]
    sheet = workbook[TABLE_OVERVIEW_SHEET]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {}, ["Table Overview was empty; table summaries use field-level data."]
    indexes = _header_indexes(header)
    if "table" not in indexes:
        return {}, ["Table Overview had no Table column; table summaries use field-level data."]
    result: dict[str, dict[str, object]] = {}
    for row_number, row in enumerate(rows, start=2):
        source_name = _bounded_text(_value(row, indexes, "table"), label=f"Table Overview row {row_number}")
        if not source_name:
            continue
        result[source_name] = {
            "description": _bounded_text(_value(row, indexes, "description"), label=f"Table {source_name} description"),
            "row_count": _optional_int(_value(row, indexes, "row_count")),
            "rows_checked": _optional_int(_value(row, indexes, "rows_checked")),
            "field_count": _optional_int(_value(row, indexes, "field_count")),
            "empty_field_count": _optional_int(_value(row, indexes, "empty_field_count")),
        }
    return result, []


def parse_whiterabbit_report(content: bytes) -> WhiteRabbitImportResult:
    """Parse one workbook without retaining value-frequency sheets."""
    _validate_archive(content)
    try:
        workbook = load_workbook(
            BytesIO(content),
            data_only=True,
            keep_links=False,
            read_only=True,
        )
    except (InvalidFileException, BadZipFile, OSError, ValueError) as error:
        raise WhiteRabbitReportError("Choose a valid .xlsx WhiteRabbit report.") from error

    try:
        if len(workbook.sheetnames) > MAXIMUM_TABLES + 10:
            raise WhiteRabbitReportError("Workbook contains too many worksheets.")
        if FIELD_OVERVIEW_SHEET not in workbook.sheetnames:
            raise WhiteRabbitReportError("Workbook is missing the Field Overview worksheet.")

        table_overview, warnings = _table_overview(workbook)
        sheet = workbook[FIELD_OVERVIEW_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise WhiteRabbitReportError("Field Overview is empty.")
        indexes = _header_indexes(header)
        missing_headers = sorted(_REQUIRED_FIELD_HEADERS - set(indexes))
        if missing_headers:
            raise WhiteRabbitReportError(
                "Field Overview is missing required columns: "
                + ", ".join(missing_headers)
                + "."
            )

        grouped: dict[str, dict[str, object]] = {}
        model_sources: dict[str, str] = {}
        field_count = 0
        renamed_table_count = 0
        renamed_field_count = 0
        for row_number, row in enumerate(rows, start=2):
            source_table = _bounded_text(
                _value(row, indexes, "table"),
                label=f"Field Overview row {row_number} table",
            )
            source_field = _bounded_text(
                _value(row, indexes, "field"),
                label=f"Field Overview row {row_number} field",
            )
            if not source_table and not source_field:
                continue
            if not source_table or not source_field:
                raise WhiteRabbitReportError(
                    f"Field Overview row {row_number} has a blank table or field."
                )
            field_count += 1
            if field_count > MAXIMUM_FIELDS:
                raise WhiteRabbitReportError("Workbook contains too many fields.")

            model_name = _normalise_identifier(
                source_table,
                label=f"Table {source_table!r}",
                maximum_characters=MAXIMUM_MODEL_IDENTIFIER_CHARACTERS,
                numeric_prefix="table",
            )
            previous_source = model_sources.setdefault(model_name, source_table)
            if previous_source != source_table:
                raise WhiteRabbitReportError(
                    f"Tables {previous_source!r} and {source_table!r} normalise to the same name."
                )
            group = grouped.setdefault(
                source_table,
                {
                    "model_name": model_name,
                    "fields": [],
                    "field_names": set(),
                    "candidate_key_fields": [],
                    "profiles": [],
                },
            )
            field_name = _normalise_identifier(
                source_field,
                label=f"Field {source_table}.{source_field}",
                maximum_characters=MAXIMUM_FIELD_IDENTIFIER_CHARACTERS,
                numeric_prefix="field",
            )
            field_names = group["field_names"]
            if field_name in field_names:
                raise WhiteRabbitReportError(
                    f"Table {source_table!r} contains duplicate normalized field {field_name!r}."
                )
            field_names.add(field_name)

            data_type, raw_data_type = _canonical_data_type(
                _value(row, indexes, "type")
            )
            row_count = _nonnegative_int(_value(row, indexes, "row_count"))
            rows_checked = _nonnegative_int(_value(row, indexes, "rows_checked"))
            fraction_empty = _fraction(_value(row, indexes, "fraction_empty"))
            fraction_unique = _fraction(_value(row, indexes, "fraction_unique"))
            unique_values, unique_values_is_upper_bound = _unique_values(
                _value(row, indexes, "unique_values")
            )
            description = _bounded_text(
                _value(row, indexes, "description"),
                label=f"Field {source_table}.{source_field} description",
            )
            if (
                rows_checked is not None
                and rows_checked > 0
                and fraction_empty == 0
                and fraction_unique == 1
            ):
                group["candidate_key_fields"].append(field_name)
            group["fields"].append(
                {
                    "name": field_name,
                    "data_type": data_type,
                    "description": description,
                }
            )
            group["profiles"].append(
                WhiteRabbitFieldProfile(
                    name=field_name,
                    source_field_name=source_field,
                    raw_data_type=raw_data_type,
                    max_length=_nonnegative_int(_value(row, indexes, "max_length")),
                    row_count=row_count,
                    rows_checked=rows_checked,
                    fraction_empty=fraction_empty,
                    unique_values=unique_values,
                    unique_values_is_upper_bound=unique_values_is_upper_bound,
                    fraction_unique=fraction_unique,
                    candidate_key=(
                        rows_checked is not None
                        and rows_checked > 0
                        and fraction_empty == 0
                        and fraction_unique == 1
                    ),
                )
            )
            if source_field.lower() != field_name:
                renamed_field_count += 1

        if not grouped:
            raise WhiteRabbitReportError("Field Overview contains no source fields.")
        if len(grouped) > MAXIMUM_TABLES:
            raise WhiteRabbitReportError("Workbook contains too many tables.")

        generated: list[WhiteRabbitGeneratedSchema] = []
        generated_content_bytes = 0
        for source_table, group in sorted(
            grouped.items(), key=lambda item: str(item[1]["model_name"])
        ):
            model_name = str(group["model_name"])
            fields = list(group["fields"])
            summary = table_overview.get(source_table, {})
            description = str(summary.get("description") or "")
            document = {
                "version": 2,
                "models": [
                    {
                        "name": model_name,
                        "description": description,
                        "columns": fields,
                    }
                ],
            }
            SourceSchemaDocument.model_validate(document)
            yaml_content = yaml.safe_dump(
                document,
                sort_keys=False,
                allow_unicode=False,
                width=100,
            )
            if len(yaml_content.encode("utf-8")) > MAXIMUM_GENERATED_YAML_BYTES:
                raise WhiteRabbitReportError(
                    f"Generated schema for {model_name!r} exceeds the document limit."
                )
            generated_content_bytes += len(yaml_content.encode("utf-8"))
            if generated_content_bytes > MAXIMUM_GENERATED_RESPONSE_BYTES:
                raise WhiteRabbitReportError(
                    "Generated source schemas exceed the import response limit."
                )
            if source_table.lower() != model_name:
                renamed_table_count += 1
            generated.append(
                WhiteRabbitGeneratedSchema(
                    file_name=f"{model_name}.yml",
                    content=yaml_content,
                    source_table_name=source_table,
                    model_name=model_name,
                    field_count=len(fields),
                    candidate_key_fields=list(group["candidate_key_fields"]),
                    profile=WhiteRabbitTableProfile(
                        model_name=model_name,
                        source_table_name=source_table,
                        description=description,
                        row_count=_nonnegative_int(summary.get("row_count")),
                        rows_checked=_nonnegative_int(summary.get("rows_checked")),
                        field_count=len(fields),
                        empty_field_count=_nonnegative_int(
                            summary.get("empty_field_count")
                        ),
                        fields=list(group["profiles"]),
                    ),
                )
            )

        if renamed_table_count:
            warnings.append(
                f"{renamed_table_count} table name(s) required structural "
                "normalization; review the generated names before importing."
            )
        if renamed_field_count:
            warnings.append(
                f"{renamed_field_count} field name(s) required structural "
                "normalization; review the generated YAML before mapping."
            )
        if any(table.candidate_key_fields for table in generated):
            warnings.append(
                "Candidate keys are statistical hints only and were not "
                "declared as primary keys."
            )
        result = WhiteRabbitImportResult(
            report_version=_report_version(workbook),
            table_count=len(generated),
            field_count=field_count,
            renamed_table_count=renamed_table_count,
            renamed_field_count=renamed_field_count,
            tables=generated,
            warnings=warnings,
        )
        if len(result.model_dump_json().encode("utf-8")) > MAXIMUM_GENERATED_RESPONSE_BYTES:
            raise WhiteRabbitReportError(
                "Generated source schemas exceed the import response limit."
            )
        return result
    finally:
        workbook.close()
