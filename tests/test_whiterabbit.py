import unittest
from io import BytesIO
from unittest.mock import patch

import yaml
from openpyxl import Workbook, load_workbook

from agent.contracts import SourceSchemaDocument
from agent.whiterabbit import (
    WhiteRabbitReportError,
    parse_whiterabbit_report,
)


FIELD_HEADERS = [
    "Table",
    "Field",
    "Description",
    "Type",
    "Max length",
    "N rows",
    "N rows checked",
    "Fraction empty",
    "N unique values",
    "Fraction unique",
]


def workbook_bytes(
    field_rows: list[list[object]],
    *,
    include_metadata: bool = True,
    include_table_overview: bool = True,
    headers: list[str] | None = None,
) -> bytes:
    """Build a compact, de-identified WhiteRabbit-compatible fixture."""
    workbook = Workbook()
    fields = workbook.active
    fields.title = "Field Overview"
    fields.append(headers or FIELD_HEADERS)
    for row in field_rows:
        fields.append(row)

    if include_table_overview:
        tables = workbook.create_sheet("Table Overview")
        tables.append(
            [
                "Table",
                "Description",
                "N rows",
                "N rows checked",
                "N Fields",
                "N Fields Empty",
            ]
        )
        source_tables = sorted({str(row[0]) for row in field_rows})
        for table in source_tables:
            tables.append([table, "Fixture table", 10, 10, 2, 0])

    if include_metadata:
        metadata = workbook.create_sheet("_")
        metadata.append(["Key", "Value"])
        metadata.append(["Version", "1.2.3"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def build_scan_report() -> bytes:
    """Build the shared API fixture with a value-frequency privacy sentinel."""
    content = workbook_bytes(
        [
            [
                "CLINICAL_PATIENT",
                "PATIENT_ID",
                "Patient identifier",
                "NUMBER",
                16,
                10,
                10,
                0,
                10,
                1,
            ]
        ]
    )
    # Reopen the compact workbook so the fixture also contains the raw-value
    # sheet that the importer must ignore entirely.
    parsed = load_workbook(BytesIO(content))
    values = parsed.create_sheet("CLINICAL_PATIENT")
    values.append(["PATIENT_ID", "Frequency"])
    values.append(["SECRET-PATIENT-VALUE", 1])
    output = BytesIO()
    parsed.save(output)
    parsed.close()
    return output.getvalue()


class WhiteRabbitImportTest(unittest.TestCase):
    def test_generates_contract_valid_yaml_and_safe_preview_metadata(self):
        content = workbook_bytes(
            [
                [
                    "Patient Data",
                    "PATIENT ID",
                    "Stable patient identifier",
                    "NUMBER",
                    16,
                    10,
                    10,
                    0,
                    10,
                    1,
                ],
                [
                    "Patient Data",
                    "DOB",
                    "Date of birth",
                    "DATE",
                    None,
                    10,
                    10,
                    0.1,
                    9,
                    0.9,
                ],
            ]
        )

        result = parse_whiterabbit_report(content)

        self.assertEqual(result.report_version, "1.2.3")
        self.assertEqual(result.table_count, 1)
        self.assertEqual(result.field_count, 2)
        self.assertEqual(result.renamed_table_count, 1)
        self.assertEqual(result.renamed_field_count, 1)
        generated = result.tables[0]
        self.assertEqual(generated.file_name, "patient_data.yml")
        self.assertEqual(generated.source_table_name, "Patient Data")
        self.assertEqual(generated.candidate_key_fields, ["patient_id"])
        self.assertEqual(generated.profile.row_count, 10)
        self.assertEqual(generated.profile.fields[0].unique_values, 10)
        self.assertTrue(generated.profile.fields[0].candidate_key)
        document = SourceSchemaDocument.model_validate(
            yaml.safe_load(generated.content)
        )
        self.assertEqual(document.models[0].name, "patient_data")
        self.assertEqual(
            [column.data_type for column in document.models[0].columns],
            ["numeric", "date"],
        )
        response_text = result.model_dump_json()
        self.assertNotIn("minimum", response_text)
        self.assertNotIn("maximum", response_text)

    def test_prefixes_identifiers_that_start_with_digits(self):
        result = parse_whiterabbit_report(
            workbook_bytes(
                [["10_PATIENT", "1_ID", "", "NUMBER", 16, 1, 1, 0, 1, 1]]
            )
        )

        self.assertEqual(result.tables[0].file_name, "table_10_patient.yml")
        document = yaml.safe_load(result.tables[0].content)
        self.assertEqual(
            document["models"][0]["columns"][0]["name"],
            "field_1_id",
        )

    def test_accepts_report_without_optional_profile_columns(self):
        result = parse_whiterabbit_report(
            workbook_bytes(
                [["PATIENT", "PATIENT_ID", "NUMBER"]],
                headers=["Table", "Field", "Type"],
                include_table_overview=False,
            )
        )

        self.assertEqual(result.table_count, 1)
        self.assertEqual(result.tables[0].candidate_key_fields, [])
        self.assertIn("Table Overview was not present", result.warnings[0])

    def test_rejects_duplicate_normalized_fields(self):
        content = workbook_bytes(
            [
                ["PATIENT", "Patient ID", "", "NUMBER", 16, 1, 1, 0, 1, 1],
                ["PATIENT", "Patient-ID", "", "NUMBER", 16, 1, 1, 0, 1, 1],
            ]
        )

        with self.assertRaisesRegex(
            WhiteRabbitReportError,
            "duplicate normalized field",
        ):
            parse_whiterabbit_report(content)

    def test_rejects_colliding_normalized_table_names(self):
        content = workbook_bytes(
            [
                ["Patient Data", "ID", "", "NUMBER", 16, 1, 1, 0, 1, 1],
                ["Patient-Data", "ID", "", "NUMBER", 16, 1, 1, 0, 1, 1],
            ]
        )

        with self.assertRaisesRegex(
            WhiteRabbitReportError,
            "normalise to the same name",
        ):
            parse_whiterabbit_report(content)

    def test_rejects_missing_required_headers(self):
        content = workbook_bytes(
            [["PATIENT", "PATIENT_ID"]],
            headers=["Table", "Field"],
        )

        with self.assertRaisesRegex(
            WhiteRabbitReportError,
            "missing required columns: type",
        ):
            parse_whiterabbit_report(content)

    def test_rejects_non_xlsx_content(self):
        with self.assertRaisesRegex(WhiteRabbitReportError, "valid .xlsx"):
            parse_whiterabbit_report(b"not an xlsx workbook")

    def test_rejects_an_oversized_generated_response(self):
        content = workbook_bytes(
            [["PATIENT", "PATIENT_ID", "", "NUMBER", 16, 1, 1, 0, 1, 1]]
        )

        with (
            patch("agent.whiterabbit.MAXIMUM_GENERATED_RESPONSE_BYTES", 1),
            self.assertRaisesRegex(
                WhiteRabbitReportError,
                "import response limit",
            ),
        ):
            parse_whiterabbit_report(content)


if __name__ == "__main__":
    unittest.main()
